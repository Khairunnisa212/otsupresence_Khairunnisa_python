from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Count
from django.core.paginator import Paginator
import csv
import io
import functools
from datetime import date, timedelta

from .models import Employee
from .forms import EmployeeForm
from attendance.models import Attendance
from accounts.models import User


import functools


def admin_required(view_func):
    @functools.wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_admin_role:
            messages.error(request, 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
def dashboard(request):
    today = date.today()
    user = request.user

    if user.is_admin_role:
        total_karyawan = Employee.objects.filter(status_aktif='aktif').count()
        hadir_hari_ini = Attendance.objects.filter(tanggal=today).count()
        terlambat = Attendance.objects.filter(tanggal=today, status='terlambat').count()
        tidak_hadir = total_karyawan - hadir_hari_ini

        recent_attendance = Attendance.objects.filter(
            tanggal=today
        ).select_related('employee').order_by('-jam_masuk')[:10]

        karyawan_tidak_hadir = Employee.objects.filter(
            status_aktif='aktif'
        ).exclude(
            attendance__tanggal=today
        )[:5]

        # Monthly stats for chart
        monthly_data = []
        for i in range(8, 0, -1):
            # Accurately go back i-1 months from current month
            month_offset = today.month - (i - 1)
            year_offset = today.year + (month_offset - 1) // 12
            month_offset = ((month_offset - 1) % 12) + 1
            month_date = date(year_offset, month_offset, 1)
            count = Attendance.objects.filter(
                tanggal__year=month_date.year,
                tanggal__month=month_date.month
            ).values('employee').distinct().count()
            monthly_data.append({
                'month': month_date.strftime('%b'),
                'count': count
            })

        context = {
            'total_karyawan': total_karyawan,
            'total_employees': total_karyawan,
            'hadir_hari_ini': hadir_hari_ini,
            'hadir_today': hadir_hari_ini,
            'terlambat': terlambat,
            'terlambat_today': terlambat,
            'tidak_hadir': tidak_hadir,
            'tidak_hadir_today': tidak_hadir,
            'recent_attendance': recent_attendance,
            'karyawan_tidak_hadir': karyawan_tidak_hadir,
            'monthly_data': monthly_data,
            'today': today,
        }
    else:
        # Employee dashboard
        try:
            employee = user.employee_profile
        except Employee.DoesNotExist:
            employee = None

        today_attendance = None
        if employee:
            today_attendance = Attendance.objects.filter(
                employee=employee, tanggal=today
            ).first()

        recent_attendance = []
        monthly_stats = {'hadir': 0, 'terlambat': 0, 'total': 0}
        total_hadir = 0
        tepat_waktu = 0
        terlambat_count = 0

        if employee:
            recent_attendance = Attendance.objects.filter(
                employee=employee
            ).order_by('-tanggal', '-jam_masuk')[:10]

            this_month = Attendance.objects.filter(
                employee=employee,
                tanggal__year=today.year,
                tanggal__month=today.month
            )
            monthly_stats['total'] = this_month.count()
            monthly_stats['hadir'] = this_month.filter(status='hadir').count()
            monthly_stats['terlambat'] = this_month.filter(status='terlambat').count()

            all_att = Attendance.objects.filter(employee=employee)
            total_hadir = all_att.count()
            tepat_waktu = all_att.filter(status='hadir').count()
            terlambat_count = all_att.filter(status='terlambat').count()

        context = {
            'employee': employee,
            'today_attendance': today_attendance,
            'recent_attendance': recent_attendance,
            'monthly_stats': monthly_stats,
            'total_hadir': total_hadir,
            'tepat_waktu': tepat_waktu,
            'terlambat_count': terlambat_count,
            'today': today,
        }

    return render(request, 'employees/dashboard.html', context)


@admin_required
def employee_list(request):
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    departemen = request.GET.get('departemen', '')
    sort = request.GET.get('sort', '-created_at')
    page = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 15)

    employees = Employee.objects.all()

    if search:
        employees = employees.filter(
            Q(nama__icontains=search) |
            Q(email__icontains=search) |
            Q(jabatan__icontains=search) |
            Q(nomor_karyawan__icontains=search)
        )
    if status:
        employees = employees.filter(status_aktif=status)
    if departemen:
        employees = employees.filter(departemen__icontains=departemen)

    valid_sorts = ['nama', '-nama', 'tanggal_masuk', '-tanggal_masuk', 'jabatan', '-created_at']
    if sort in valid_sorts:
        employees = employees.order_by(sort)

    paginator = Paginator(employees, int(per_page))
    page_obj = paginator.get_page(page)

    departemen_list = Employee.objects.values_list('departemen', flat=True).distinct()

    context = {
        'page_obj': page_obj,
        'search': search,
        'status': status,
        'departemen': departemen,
        'sort': sort,
        'per_page': per_page,
        'departemen_list': departemen_list,
        'total_count': employees.count(),
    }
    return render(request, 'employees/employee_list.html', context)


@admin_required
def employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            employee = form.save()
            # Create user account for employee
            user_data = form.cleaned_data
            if user_data.get('create_account'):
                user = User.objects.create_user(
                    username=user_data.get('username', employee.email.split('@')[0]),
                    email=employee.email,
                    password=user_data.get('password', 'karyawan123'),
                    first_name=employee.nama.split()[0],
                    last_name=' '.join(employee.nama.split()[1:]),
                    role=User.ROLE_EMPLOYEE,
                )
                employee.user = user
                employee.save()
            messages.success(request, f'Karyawan {employee.nama} berhasil ditambahkan!')
            return redirect('employee_list')
    else:
        form = EmployeeForm()
    return render(request, 'employees/employee_form.html', {'form': form, 'action': 'Tambah'})


@admin_required
def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, f'Data {employee.nama} berhasil diperbarui!')
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'employees/employee_form.html', {'form': form, 'action': 'Edit', 'employee': employee})


@admin_required
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        nama = employee.nama
        employee.delete()
        messages.success(request, f'Karyawan {nama} berhasil dihapus!')
        return redirect('employee_list')
    return render(request, 'employees/employee_confirm_delete.html', {'employee': employee})


@admin_required
def employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    today = date.today()

    attendance_history = Attendance.objects.filter(
        employee=employee
    ).order_by('-tanggal')[:30]

    monthly_stats = {}
    for i in range(6):
        month_offset = today.month - i
        year_offset = today.year + (month_offset - 1) // 12
        month_offset = ((month_offset - 1) % 12) + 1
        month_date = date(year_offset, month_offset, 1)
        month_att = Attendance.objects.filter(
            employee=employee,
            tanggal__year=month_date.year,
            tanggal__month=month_date.month
        )
        monthly_stats[month_date.strftime('%B %Y')] = {
            'hadir': month_att.filter(status='hadir').count(),
            'terlambat': month_att.filter(status='terlambat').count(),
            'total': month_att.count(),
        }

    context = {
        'employee': employee,
        'attendance_history': attendance_history,
        'monthly_stats': monthly_stats,
    }
    return render(request, 'employees/employee_detail.html', context)


@admin_required
def export_employees_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data Karyawan'

    # Header styling
    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ['No', 'Nomor Karyawan', 'Nama', 'Email', 'Jabatan', 'Departemen', 'Tanggal Masuk', 'Status']
    col_widths = [5, 18, 25, 30, 20, 20, 18, 12]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 25

    employees = Employee.objects.all().order_by('nama')
    for i, emp in enumerate(employees, 1):
        row_data = [
            i, emp.nomor_karyawan, emp.nama, emp.email,
            emp.jabatan, emp.departemen,
            emp.tanggal_masuk.strftime('%d/%m/%Y'),
            emp.get_status_aktif_display()
        ]
        fill_color = 'F5F9FF' if i % 2 == 0 else 'FFFFFF'
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.border = border
            cell.fill = row_fill
            cell.alignment = Alignment(vertical='center')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="data_karyawan_{date.today()}.xlsx"'
    return response


@admin_required
def export_employees_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="data_karyawan_{date.today()}.csv"'

    writer = csv.writer(response)
    writer.writerow(['No', 'Nomor Karyawan', 'Nama', 'Email', 'Jabatan', 'Departemen', 'Tanggal Masuk', 'Status'])

    for i, emp in enumerate(Employee.objects.all().order_by('nama'), 1):
        writer.writerow([
            i, emp.nomor_karyawan, emp.nama, emp.email,
            emp.jabatan, emp.departemen,
            emp.tanggal_masuk.strftime('%d/%m/%Y'),
            emp.get_status_aktif_display()
        ])

    return response


@admin_required
def export_employees_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="data_karyawan_{date.today()}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, textColor=colors.HexColor('#1565C0'), spaceAfter=20)
    elements.append(Paragraph('Laporan Data Karyawan', title_style))
    elements.append(Paragraph(f'Tanggal: {date.today().strftime("%d %B %Y")}', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    employees = Employee.objects.all().order_by('nama')
    data = [['No', 'Nomor', 'Nama', 'Email', 'Jabatan', 'Departemen', 'Tgl Masuk', 'Status']]
    for i, emp in enumerate(employees, 1):
        data.append([
            str(i), emp.nomor_karyawan, emp.nama, emp.email,
            emp.jabatan, emp.departemen,
            emp.tanggal_masuk.strftime('%d/%m/%Y'),
            emp.get_status_aktif_display()
        ])

    table = Table(data, colWidths=[1*cm, 2.5*cm, 4*cm, 5*cm, 3.5*cm, 3*cm, 2.5*cm, 2*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F9FF')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWHEIGHT', (0, 0), (-1, -1), 20),
    ]))
    elements.append(table)
    doc.build(elements)
    return response
