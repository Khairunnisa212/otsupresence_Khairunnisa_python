from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator
from datetime import date, time, timedelta
import csv
import io

from .models import Attendance
from employees.models import Employee
from employees.views import admin_required


@login_required
def check_in(request):
    if request.method != 'POST':
        return redirect('dashboard')
    user = request.user
    try:
        employee = user.employee_profile
    except Exception:
        messages.error(request, 'Profil karyawan tidak ditemukan.')
        return redirect('dashboard')

    today = date.today()
    existing = Attendance.objects.filter(employee=employee, tanggal=today).first()

    if existing:
        messages.warning(request, 'Anda sudah melakukan check-in hari ini.')
        return redirect('dashboard')

    now = timezone.localtime(timezone.now()).time()
    
    attendance = Attendance(
        employee=employee,
        tanggal=today,
        jam_masuk=now,
        lokasi_masuk=request.POST.get('lokasi', 'Kantor Pusat'),
    )
    
    # Handle GPS coordinates
    lat = request.POST.get('latitude')
    lng = request.POST.get('longitude')
    if lat and lng:
        try:
            attendance.latitude_masuk = float(lat)
            attendance.longitude_masuk = float(lng)
        except (ValueError, TypeError):
            pass
    
    # Handle photo upload
    foto = request.FILES.get('foto_masuk')
    if foto:
        attendance.foto_masuk = foto
    
    attendance.save()
    messages.success(request, f'Check-in berhasil pukul {now.strftime("%H:%M")}! {"Anda terlambat." if attendance.status == "terlambat" else "Tepat waktu!"}')
    return redirect('dashboard')


@login_required
def check_out(request):
    if request.method != 'POST':
        return redirect('dashboard')
    user = request.user
    try:
        employee = user.employee_profile
    except Exception:
        messages.error(request, 'Profil karyawan tidak ditemukan.')
        return redirect('dashboard')

    today = date.today()
    attendance = Attendance.objects.filter(employee=employee, tanggal=today).first()

    if not attendance:
        messages.error(request, 'Anda belum melakukan check-in hari ini.')
        return redirect('dashboard')

    if attendance.jam_keluar:
        messages.warning(request, 'Anda sudah melakukan check-out hari ini.')
        return redirect('dashboard')

    now = timezone.localtime(timezone.now()).time()
    attendance.jam_keluar = now
    attendance.lokasi_keluar = request.POST.get('lokasi', '')
    
    lat = request.POST.get('latitude')
    lng = request.POST.get('longitude')
    if lat and lng:
        try:
            attendance.latitude_keluar = float(lat)
            attendance.longitude_keluar = float(lng)
        except (ValueError, TypeError):
            pass
    
    foto = request.FILES.get('foto_keluar')
    if foto:
        attendance.foto_keluar = foto
    
    attendance.save()
    messages.success(request, f'Check-out berhasil pukul {now.strftime("%H:%M")}! Durasi kerja: {attendance.durasi_kerja}')
    return redirect('dashboard')


@login_required
def attendance_list(request):
    user = request.user
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    page = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 15)

    if user.is_admin_role:
        attendances = Attendance.objects.select_related('employee').all()
    else:
        try:
            employee = user.employee_profile
            attendances = Attendance.objects.filter(employee=employee)
        except Exception:
            attendances = Attendance.objects.none()

    if search and user.is_admin_role:
        attendances = attendances.filter(
            Q(employee__nama__icontains=search) |
            Q(employee__nomor_karyawan__icontains=search)
        )
    if status:
        attendances = attendances.filter(status=status)
    if date_from:
        attendances = attendances.filter(tanggal__gte=date_from)
    if date_to:
        attendances = attendances.filter(tanggal__lte=date_to)

    attendances = attendances.order_by('-tanggal', '-jam_masuk')
    paginator = Paginator(attendances, int(per_page))
    page_obj = paginator.get_page(page)

    context = {
        'page_obj': page_obj,
        'search': search,
        'status': status,
        'date_from': date_from,
        'date_to': date_to,
        'per_page': per_page,
        'total_count': attendances.count(),
        'is_admin': user.is_admin_role,
    }
    return render(request, 'attendance/attendance_list.html', context)


@login_required
def monthly_report(request):
    user = request.user
    today = date.today()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))

    if user.is_admin_role:
        employees = Employee.objects.filter(status_aktif='aktif')
        report_data = []
        for emp in employees:
            att = Attendance.objects.filter(
                employee=emp, tanggal__year=year, tanggal__month=month
            )
            report_data.append({
                'employee': emp,
                'hadir': att.filter(status='hadir').count(),
                'terlambat': att.filter(status='terlambat').count(),
                'total': att.count(),
            })
    else:
        try:
            employee = user.employee_profile
        except Exception:
            return redirect('dashboard')
        att = Attendance.objects.filter(
            employee=employee, tanggal__year=year, tanggal__month=month
        ).order_by('tanggal')
        report_data = [{
            'employee': employee,
            'attendance_list': att,
            'hadir': att.filter(status='hadir').count(),
            'terlambat': att.filter(status='terlambat').count(),
            'total': att.count(),
        }]

    months = [(i, date(2024, i, 1).strftime('%B')) for i in range(1, 13)]
    years = list(range(2023, today.year + 2))

    # Build monthly chart data for admin
    monthly_data = []
    monthly_max = 1
    if user.is_admin_role:
        from employees.views import dashboard as _dash
        for i in range(8, 0, -1):
            mo = today.month - (i - 1)
            yr = today.year + (mo - 1) // 12
            mo = ((mo - 1) % 12) + 1
            md = date(yr, mo, 1)
            cnt = Attendance.objects.filter(tanggal__year=md.year, tanggal__month=md.month).count()
            monthly_data.append({"month": md.strftime("%b"), "count": cnt})
        monthly_max = max((d["count"] for d in monthly_data), default=1) or 1

    context = {
        'report_data': report_data,
        'year': year,
        'month': month,
        'months': months,
        'years': years,
        'is_admin': user.is_admin_role,
        'monthly_data': monthly_data,
        'monthly_max': monthly_max,
    }
    return render(request, 'attendance/monthly_report.html', context)


@admin_required
def export_attendance_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    year = request.GET.get('year', date.today().year)
    month = request.GET.get('month', date.today().month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Absensi'

    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['No', 'Nomor Karyawan', 'Nama', 'Tanggal', 'Jam Masuk', 'Jam Keluar', 'Durasi', 'Status']
    col_widths = [5, 18, 25, 15, 12, 12, 12, 12]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    attendances = Attendance.objects.filter(
        tanggal__year=year, tanggal__month=month
    ).select_related('employee').order_by('tanggal', 'employee__nama')

    for i, att in enumerate(attendances, 1):
        fill_color = 'F5F9FF' if i % 2 == 0 else 'FFFFFF'
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        row_data = [
            i, att.employee.nomor_karyawan, att.employee.nama,
            att.tanggal.strftime('%d/%m/%Y'),
            att.jam_masuk.strftime('%H:%M') if att.jam_masuk else '-',
            att.jam_keluar.strftime('%H:%M') if att.jam_keluar else '-',
            att.durasi_kerja,
            att.get_status_display()
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.border = border
            cell.fill = row_fill
            cell.alignment = Alignment(vertical='center')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="absensi_{year}_{month}.xlsx"'
    return response


@admin_required
def export_attendance_csv(request):
    year = request.GET.get('year', date.today().year)
    month = request.GET.get('month', date.today().month)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="absensi_{year}_{month}.csv"'

    writer = csv.writer(response)
    writer.writerow(['No', 'Nomor Karyawan', 'Nama', 'Tanggal', 'Jam Masuk', 'Jam Keluar', 'Durasi', 'Status'])

    attendances = Attendance.objects.filter(
        tanggal__year=year, tanggal__month=month
    ).select_related('employee').order_by('tanggal', 'employee__nama')

    for i, att in enumerate(attendances, 1):
        writer.writerow([
            i, att.employee.nomor_karyawan, att.employee.nama,
            att.tanggal.strftime('%d/%m/%Y'),
            att.jam_masuk.strftime('%H:%M') if att.jam_masuk else '-',
            att.jam_keluar.strftime('%H:%M') if att.jam_keluar else '-',
            att.durasi_kerja,
            att.get_status_display()
        ])
    return response


@admin_required
def export_attendance_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    year = request.GET.get('year', date.today().year)
    month = request.GET.get('month', date.today().month)
    month_name = date(int(year), int(month), 1).strftime('%B %Y')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="absensi_{year}_{month}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16,
                                  textColor=colors.HexColor('#1565C0'), spaceAfter=10)
    elements.append(Paragraph(f'Laporan Absensi - {month_name}', title_style))
    elements.append(Spacer(1, 0.5*cm))

    attendances = Attendance.objects.filter(
        tanggal__year=year, tanggal__month=month
    ).select_related('employee').order_by('tanggal', 'employee__nama')

    data = [['No', 'Nomor', 'Nama', 'Tanggal', 'Jam Masuk', 'Jam Keluar', 'Durasi', 'Status']]
    for i, att in enumerate(attendances, 1):
        data.append([
            str(i), att.employee.nomor_karyawan, att.employee.nama,
            att.tanggal.strftime('%d/%m/%Y'),
            att.jam_masuk.strftime('%H:%M') if att.jam_masuk else '-',
            att.jam_keluar.strftime('%H:%M') if att.jam_keluar else '-',
            att.durasi_kerja,
            att.get_status_display()
        ])

    table = Table(data, colWidths=[1*cm, 2.5*cm, 4.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm, 2*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F9FF')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table)
    doc.build(elements)
    return response
